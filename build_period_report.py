# build_period_report.py
import os
import argparse
import pandas as pd
import firebirdsql
from openpyxl.utils import get_column_letter
import config

pd.set_option("display.max_columns", None)
pd.set_option("display.width", 1600)

OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Ordem de tentativas de charset
CHARSETS = []
if getattr(config, "CHARSET", None):
    CHARSETS.append(config.CHARSET)
for cs in ["UTF8", "WIN1252", "ISO8859_1", "DOS850"]:
    if cs not in CHARSETS:
        CHARSETS.append(cs)

# ----------------- CLI -----------------
def parse_args():
    p = argparse.ArgumentParser(
        description="Relatório único (Detalhe/Resumo/Vendas do Período/Tipos de Pagto/Contadores/Resumo por Vendedor) + CSVs"
    )
    p.add_argument("--start", dest="start", help="YYYY-MM-DD (inclusivo)")
    p.add_argument("--end", dest="end", help="YYYY-MM-DD (inclusivo)")
    return p.parse_args()

# ----------------- DB helpers -----------------
def connect_with(cs: str):
    return firebirdsql.connect(
        host=config.DB_HOST,
        port=int(str(config.DB_PORT or "3369")),
        database=config.DB_PATH,
        user=config.DB_USER,
        password=config.DB_PASSWORD,
        charset=cs,
    )

def exec_sql(sql: str) -> pd.DataFrame:
    last = None
    for cs in CHARSETS:
        try:
            with connect_with(cs) as conn:
                return pd.read_sql(sql, conn)
        except Exception as e:
            last = e
            print(f"⚠️ Falhou charset={cs}: {e}")
    raise last  # type: ignore

def build_where_date(date_col: str, start: str | None, end: str | None) -> str:
    if start and end:
        return f"WHERE CAST({date_col} AS DATE) BETWEEN DATE '{start}' AND DATE '{end}'"
    elif start:
        return f"WHERE CAST({date_col} AS DATE) >= DATE '{start}'"
    elif end:
        return f"WHERE CAST({date_col} AS DATE) <= DATE '{end}'"
    else:
        return ""

def where_plus(base_where: str, extra_cond: str) -> str:
    return (base_where + (" AND " if base_where else " WHERE ") + extra_cond)

# ----------------- utils -----------------
def dia_semana_pt(dt_series: pd.Series) -> pd.Series:
    mapa = {0:"Segunda-feira",1:"Terça-feira",2:"Quarta-feira",
            3:"Quinta-feira",4:"Sexta-feira",5:"Sábado",6:"Domingo"}
    return dt_series.dt.weekday.map(mapa)

def safe_div(num, den):
    try:
        den = float(den)
        return 0.0 if den == 0.0 else float(num) / den
    except Exception:
        return 0.0

def excel_number_formats(ws, header_row_idx: int, df: pd.DataFrame, money_cols: list[str], int_cols: list[str]):
    for col_idx, col_name in enumerate(df.columns, start=1):
        xl_col = get_column_letter(col_idx)
        if col_name in money_cols:
            for row in range(header_row_idx + 1, header_row_idx + 1 + len(df)):
                ws[f"{xl_col}{row}"].number_format = "#,##0.00"
        elif col_name in int_cols:
            for row in range(header_row_idx + 1, header_row_idx + 1 + len(df)):
                ws[f"{xl_col}{row}"].number_format = "0"

def append_dedup(df: pd.DataFrame, path: str, key_cols: list[str]):
    """Append com dedupe pelas chaves informadas."""
    df = df.copy()
    # normaliza Data para string YYYY-MM-DD se existir
    for kc in key_cols:
        if kc.lower() == "data" and kc in df.columns:
            df[kc] = pd.to_datetime(df[kc]).dt.strftime("%Y-%m-%d")
    if os.path.exists(path):
        old = pd.read_csv(path)
        combo = pd.concat([old, df], ignore_index=True)
        combo = combo.drop_duplicates(key_cols, keep="last")
        combo.to_csv(path, index=False, encoding="utf-8-sig")
    else:
        df.to_csv(path, index=False, encoding="utf-8-sig")

# ----------------- main -----------------
if __name__ == "__main__":
    args = parse_args()
    START_DATE = args.start
    END_DATE = args.end

    # ---------- Detalhe Diário (PEDIDOS) ----------
    # [Ajuste: usar DATA_VENDA como eixo de data]
    where_ped = build_where_date("PEDIDOS.DATA_VENDA", START_DATE, END_DATE)
    SQL_daily = f"""
        SELECT
            CAST(PEDIDOS.DATA_VENDA AS DATE) AS DATA,  -- eixo: DATA_VENDA
            SUM(CASE
                    WHEN UPPER(TRIM(PEDIDOS.SITUACAO)) IN ('VENDA SEPD', 'PEDIDO DE VENDA', 'VDA HOMOLOG')
                    THEN COALESCE(PEDIDOS.VALOR_FINAL, 0)
                    ELSE 0
                END) AS TOTAL_LIQUIDO,
            SUM(CASE WHEN UPPER(TRIM(PEDIDOS.SITUACAO)) = 'VENDA SEPD'
                     THEN COALESCE(PEDIDOS.QUANT_TRANSP, 0) ELSE 0 END)
            + SUM(CASE WHEN UPPER(TRIM(PEDIDOS.SITUACAO)) = 'PEDIDO DE VENDA'
                       THEN COALESCE(PEDIDOS.PSEQ_ITEM, 0) ELSE 0 END)
            + SUM(CASE WHEN UPPER(TRIM(PEDIDOS.SITUACAO)) = 'VDA HOMOLOG'
                       THEN COALESCE(PEDIDOS.PSEQ_ITEM, 0) ELSE 0 END) AS QTDE_PRODUTOS,
            SUM(CASE WHEN UPPER(TRIM(PEDIDOS.SITUACAO)) IN ('VENDA SEPD','PEDIDO DE VENDA','VDA HOMOLOG')
                     THEN 1 ELSE 0 END) AS QTDE_PEDIDOS,
            SUM(COALESCE(PEDIDOS.DESCONTO_TOTAL, 0))  AS DESCONTO,
            SUM(COALESCE(PEDIDOS.OUTRAS_DESPESAS, 0)) AS ACRESCIMO
        FROM PEDIDOS
        {where_ped}
        GROUP BY 1
        ORDER BY 1
    """
    df = exec_sql(SQL_daily)

    # ---------- Devoluções por dia (TROCA_MERC) ----------
    # [Ajuste: alinhar eixo de data para DATA_VENDA]
    SQL_dev_day = f"""
        SELECT
            CAST(DATA_VENDA AS DATE) AS DATA,
            SUM(CASE WHEN UPPER(TRIM(SITUACAO)) = 'TROCA_MERC'
                     THEN COALESCE(VALOR_FINAL, 0) ELSE 0 END) AS DEVOLUCOES_VENDA
        FROM PEDIDOS
        {where_ped}
        GROUP BY 1
        ORDER BY 1
    """
    df_dev = exec_sql(SQL_dev_day)

    # ---------- Crédito Cliente por dia (MOVCAIXA) ----------
    # Permanece por DATA_HORA do MOVCAIXA
    where_mov = build_where_date("MOVCAIXA.DATA_HORA", START_DATE, END_DATE)
    SQL_credit_day = f"""
        SELECT
            CAST(DATA_HORA AS DATE) AS DATA,
            SUM(CASE WHEN UPPER(TRIM(DOCUMENTO)) = 'VENDA'
                     THEN COALESCE(VALOR_APRAZO, 0) ELSE 0 END) AS CREDITO_CLIENTE
        FROM MOVCAIXA
        {where_mov}
        GROUP BY 1
        ORDER BY 1
    """
    df_credit = exec_sql(SQL_credit_day)

    # Tipagem e merges
    if "DATA" in df.columns:
        df["DATA"] = pd.to_datetime(df["DATA"])
    for c in ["TOTAL_LIQUIDO", "QTDE_PRODUTOS", "QTDE_PEDIDOS", "DESCONTO", "ACRESCIMO"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    if not df_dev.empty:
        df_dev["DATA"] = pd.to_datetime(df_dev["DATA"])
        df = df.merge(df_dev, on="DATA", how="left")
    else:
        df["DEVOLUCOES_VENDA"] = 0.0

    if not df_credit.empty:
        df_credit["DATA"] = pd.to_datetime(df_credit["DATA"])
        df = df.merge(df_credit, on="DATA", how="left")
    else:
        df["CREDITO_CLIENTE"] = 0.0

    df["DEVOLUCOES_VENDA"] = pd.to_numeric(df.get("DEVOLUCOES_VENDA", 0), errors="coerce").fillna(0.0)
    df["CREDITO_CLIENTE"] = pd.to_numeric(df.get("CREDITO_CLIENTE", 0), errors="coerce").fillna(0.0)

    # Enriquecimentos do Detalhe
    df["Dia da Semana"] = dia_semana_pt(df["DATA"])
    df["Vl. Médio Produto (D)"] = df.apply(lambda r: safe_div(r["TOTAL_LIQUIDO"], r["QTDE_PRODUTOS"]), axis=1)
    df["Vl. Médio Pedido (E)"]  = df.apply(lambda r: safe_div(r["TOTAL_LIQUIDO"], r["QTDE_PEDIDOS"]), axis=1)
    df["Tot. Bruto"]            = df["TOTAL_LIQUIDO"] + df["DESCONTO"] - df["ACRESCIMO"]

    detalhe = (
        df[
            [
                "DATA",
                "Dia da Semana",
                "TOTAL_LIQUIDO",
                "QTDE_PRODUTOS",
                "QTDE_PEDIDOS",
                "Vl. Médio Produto (D)",
                "Vl. Médio Pedido (E)",
                "DESCONTO",
                "ACRESCIMO",
                "Tot. Bruto",
                "DEVOLUCOES_VENDA",
                "CREDITO_CLIENTE",
            ]
        ]
        .rename(
            columns={
                "DATA": "Data",
                "TOTAL_LIQUIDO": "Total Líquido (A)",
                "QTDE_PRODUTOS": "Qtde. Produtos (B)",
                "QTDE_PEDIDOS": "Qtde. Pedidos (C)",
                "DESCONTO": "Desconto",
                "ACRESCIMO": "Acréscimo",
                "DEVOLUCOES_VENDA": "Devoluções Venda",
                "CREDITO_CLIENTE": "Crédito Cliente",
            }
        )
        .sort_values("Data")
        .reset_index(drop=True)
    )

    # ---------- Resumo por dia da semana ----------
    ordem_semana = [
        "Segunda-feira",
        "Terça-feira",
        "Quarta-feira",
        "Quinta-feira",
        "Sexta-feira",
        "Sábado",
        "Domingo",
    ]
    detalhe["Dia da Semana"] = pd.Categorical(detalhe["Dia da Semana"], categories=ordem_semana, ordered=True)

    resumo = (
        detalhe.groupby("Dia da Semana", as_index=False)
        .agg(
            **{
                "Total Líquido (A)": ("Total Líquido (A)", "sum"),
                "Qtde. Produtos (B)": ("Qtde. Produtos (B)", "sum"),
                "Qtde. Pedidos (C)": ("Qtde. Pedidos (C)", "sum"),
                "Desconto": ("Desconto", "sum"),
                "Acréscimo": ("Acréscimo", "sum"),
                "Devoluções Venda": ("Devoluções Venda", "sum"),
                "Crédito Cliente": ("Crédito Cliente", "sum"),
            }
        )
        .sort_values("Dia da Semana")
        .reset_index(drop=True)
    )
    resumo["Vl. Médio Produto (D)"] = resumo.apply(
        lambda r: safe_div(r["Total Líquido (A)"], r["Qtde. Produtos (B)"]), axis=1
    )
    resumo["Vl. Médio Pedido (E)"] = resumo.apply(
        lambda r: safe_div(r["Total Líquido (A)"], r["Qtde. Pedidos (C)"]), axis=1
    )
    resumo["Tot. Bruto"] = resumo["Total Líquido (A)"] + resumo["Desconto"] - resumo["Acréscimo"]

    # ---------- Vendas Total do Período ----------
    # Devoluções (TROCA_MERC) baseadas em VL_BRUTO (ajuste solicitado)
    SQL_devol = f"""
        SELECT
            SUM(CASE WHEN UPPER(TRIM(SITUACAO)) = 'TROCA_MERC'
                     THEN COALESCE(VL_BRUTO, 0) ELSE 0 END)  AS DEV_BRUTO,
            SUM(CASE WHEN UPPER(TRIM(SITUACAO)) = 'TROCA_MERC'
                     THEN COALESCE(DESCONTO_TOTAL, 0) ELSE 0 END) AS DEV_DESC,
            SUM(CASE WHEN UPPER(TRIM(SITUACAO)) = 'PEDIDO DE VENDA'
                     THEN COALESCE(OUTRAS_DESPESAS, 0) ELSE 0 END) AS DEV_ACRES
        FROM PEDIDOS
        {where_ped}
    """
    SQL_devol_cnt = f"""
        SELECT COUNT(*) AS DEV_QTD
        FROM PEDIDOS
        {where_plus(where_ped, "UPPER(TRIM(SITUACAO)) = 'TROCA_MERC'")}
    """
    tot_dev = exec_sql(SQL_devol).iloc[0].fillna(0).to_dict()
    dev_cnt = int(exec_sql(SQL_devol_cnt).iloc[0]["DEV_QTD"])

    venda_bruta_total_bruto = float(detalhe["Tot. Bruto"].sum())
    venda_bruta_total_liq = float(detalhe["Total Líquido (A)"].sum())
    venda_bruta_descontos = float(detalhe["Desconto"].sum())
    venda_bruta_acrescimos = float(detalhe["Acréscimo"].sum())

    venda_bruta = {
        "Total Bruto": venda_bruta_total_bruto,
        "Descontos": venda_bruta_descontos,
        "Acréscimos": venda_bruta_acrescimos,
        "Total Líquido": venda_bruta_total_liq,
    }

    devolucoes = {
        "Total Bruto": float(tot_dev["DEV_BRUTO"]),
        "Descontos": float(tot_dev["DEV_DESC"]),
        "Acréscimos": float(tot_dev["DEV_ACRES"]),
    }
    devolucoes["Total Líquido"] = devolucoes["Total Bruto"] - devolucoes["Descontos"] + devolucoes["Acréscimos"]

    venda_liquida = {
        "Total Bruto": venda_bruta["Total Bruto"] - devolucoes["Total Bruto"],
        "Descontos": venda_bruta["Descontos"] - devolucoes["Descontos"],
        "Acréscimos": venda_bruta["Acréscimos"] - devolucoes["Acréscimos"],
        "Total Líquido": venda_bruta["Total Líquido"] - devolucoes["Total Líquido"],
    }

    venda_total_periodo = pd.DataFrame(
        {"Venda Bruta": venda_bruta, "Devoluções Venda": devolucoes, "Venda Líquida": venda_liquida}
    )

    # ---------- Totais por Tipo de Pagamento ----------
    # crédito cliente já apurado por dia para o CSV; aqui usamos o total do período:
    credito_cliente_total = float(df["CREDITO_CLIENTE"].sum())

    vista_bruta = venda_bruta_total_liq - credito_cliente_total
    vista_dev = devolucoes["Total Líquido"]
    vista_liq = vista_bruta - vista_dev

    prazo_bruta = prazo_dev = prazo_liq = 0.0  # placeholder (não há regra definida)
    credito_bruta = credito_cliente_total
    credito_dev = 0.0
    credito_liq = credito_bruta

    tipos_pagto = pd.DataFrame(
        {
            "Venda Bruta": [vista_bruta, prazo_bruta, credito_bruta],
            "Devoluções Venda": [vista_dev, prazo_dev, credito_dev],
            "Venda Líquida": [vista_liq, prazo_liq, credito_liq],
        },
        index=["Vendas à Vista", "Vendas a Prazo", "Crédito Cliente"],
    )

    # ---------- Contadores ----------
    total_pedidos = int(detalhe["Qtde. Pedidos (C)"].sum())
    total_devolucoes = dev_cnt
    contadores = pd.DataFrame(
        {"Métrica": ["Total de Pedidos", "Total de Devoluções"], "Qtde": [total_pedidos, total_devolucoes]}
    )

    # ---------- Resumo por Vendedor (PERÍODO - para o Excel) ----------
    # Vendas brutas por vendedor
    where_vend = where_plus(where_ped, "UPPER(TRIM(SITUACAO)) IN ('VENDA SEPD','PEDIDO DE VENDA','VDA HOMOLOG')")
    SQL_sellers = f"""
        SELECT
            COALESCE(NULLIF(TRIM(NOME_VENDEDOR), ''), 'Sem Vendedor') AS VENDEDOR,
            SUM(COALESCE(VALOR_FINAL, 0)) AS VENDA_BRUTA
        FROM PEDIDOS
        {where_vend}
        GROUP BY 1
    """
    vendedores = exec_sql(SQL_sellers)

    # Devoluções por vendedor
    where_dev_vend = where_plus(where_ped, "UPPER(TRIM(SITUACAO)) = 'TROCA_MERC'")
    SQL_dev_sellers = f"""
        SELECT
            COALESCE(NULLIF(TRIM(NOME_VENDEDOR), ''), 'Sem Vendedor') AS VENDEDOR,
            SUM(COALESCE(VALOR_FINAL, 0)) AS DEVOLUCOES
        FROM PEDIDOS
        {where_dev_vend}
        GROUP BY 1
    """
    vendedores_dev = exec_sql(SQL_dev_sellers)

    # Merge e cálculo da Venda Líquida
    if not vendedores.empty:
        if not vendedores_dev.empty:
            vendedores = vendedores.merge(vendedores_dev, on="VENDEDOR", how="left")
            vendedores["DEVOLUCOES"] = vendedores["DEVOLUCOES"].fillna(0)
        else:
            vendedores["DEVOLUCOES"] = 0
        
        vendedores["VENDA_LIQUIDA"] = vendedores["VENDA_BRUTA"] - vendedores["DEVOLUCOES"]
        vendedores = vendedores[["VENDEDOR", "VENDA_LIQUIDA"]].rename(
            columns={"VENDEDOR": "Vendedor", "VENDA_LIQUIDA": "Venda Líquida"}
        ).sort_values("Venda Líquida", ascending=False).reset_index(drop=True)
    else:
        vendedores = pd.DataFrame(columns=["Vendedor", "Venda Líquida"])

    # ---------- Fato diário por VENDEDOR (para CSV) ----------
    # CORREÇÃO: buscar vendas E devoluções por (Data, Vendedor)
    where_vend_day = where_plus(where_ped, "UPPER(TRIM(SITUACAO)) IN ('VENDA SEPD','PEDIDO DE VENDA','VDA HOMOLOG')")
    SQL_vendor_daily = f"""
        SELECT
            CAST(DATA_VENDA AS DATE) AS DATA,
            COALESCE(NULLIF(TRIM(NOME_VENDEDOR), ''), 'Sem Vendedor') AS VENDEDOR,
            SUM(COALESCE(VALOR_FINAL, 0)) AS VENDA_BRUTA
        FROM PEDIDOS
        {where_vend_day}
        GROUP BY 1, 2
        ORDER BY 1, 2
    """
    vendor_daily = exec_sql(SQL_vendor_daily)
    
    # Devoluções por (Data, Vendedor)
    where_dev_vend_day = where_plus(where_ped, "UPPER(TRIM(SITUACAO)) = 'TROCA_MERC'")
    SQL_vendor_dev_daily = f"""
        SELECT
            CAST(DATA_VENDA AS DATE) AS DATA,
            COALESCE(NULLIF(TRIM(NOME_VENDEDOR), ''), 'Sem Vendedor') AS VENDEDOR,
            SUM(COALESCE(VALOR_FINAL, 0)) AS DEVOLUCOES
        FROM PEDIDOS
        {where_dev_vend_day}
        GROUP BY 1, 2
        ORDER BY 1, 2
    """
    vendor_dev_daily = exec_sql(SQL_vendor_dev_daily)
    
    if not vendor_daily.empty:
        # Merge vendas com devoluções
        if not vendor_dev_daily.empty:
            vendor_daily = vendor_daily.merge(vendor_dev_daily, on=["DATA", "VENDEDOR"], how="left")
            vendor_daily["DEVOLUCOES"] = vendor_daily["DEVOLUCOES"].fillna(0)
        else:
            vendor_daily["DEVOLUCOES"] = 0
        
        # Calcula Venda Líquida
        vendor_daily["VENDA_LIQUIDA"] = vendor_daily["VENDA_BRUTA"] - vendor_daily["DEVOLUCOES"]
        
        # Renomeia e seleciona colunas
        vendor_daily = vendor_daily[["DATA", "VENDEDOR", "VENDA_LIQUIDA"]].rename(
            columns={"DATA": "Data", "VENDEDOR": "Vendedor", "VENDA_LIQUIDA": "Venda Líquida"}
        )
        vendor_daily["Data"] = pd.to_datetime(vendor_daily["Data"]).dt.strftime("%Y-%m-%d")
        out_vendor = os.path.join(OUTPUT_DIR, "fato_vendas_vendedor_diario.csv")
        append_dedup(vendor_daily, out_vendor, key_cols=["Data", "Vendedor"])
        print(f"🟢 Atualizado: {out_vendor}")
    else:
        print("ℹ️ Nenhuma linha para fato_vendas_vendedor_diario no período.")

    # ---------- Escrita do CSV diário principal (para o Looker) ----------
    fato_daily = detalhe.copy()
    # ordena colunas principais (mantendo as métricas novas)
    cols_order = [
        "Data",
        "Dia da Semana",
        "Total Líquido (A)",
        "Qtde. Produtos (B)",
        "Qtde. Pedidos (C)",
        "Desconto",
        "Acréscimo",
        "Tot. Bruto",
        "Devoluções Venda",
        "Crédito Cliente",
    ]
    fato_daily = fato_daily[cols_order]
    out_fact = os.path.join(OUTPUT_DIR, "fato_vendas_diario.csv")
    append_dedup(fato_daily, out_fact, key_cols=["Data"])
    print(f"🟢 Atualizado: {out_fact}")

    # ---------- Escrita em UMA ÚNICA ABA (Excel) ----------
    def fname_suffix():
        if START_DATE and END_DATE:
            return f"_{START_DATE.replace('-','')}_{END_DATE.replace('-','')}"
        if START_DATE:
            return f"_{START_DATE.replace('-','')}_"
        if END_DATE:
            return f"__{END_DATE.replace('-','')}"
        return ""

    xlsx_path = os.path.join(OUTPUT_DIR, f"RelatorioPeriodo{fname_suffix()}.xlsx")
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        sheet = "RelatorioPeriodo"

        # Parâmetros
        params = pd.DataFrame(
            [
                {"Parâmetro": "Data Inicial", "Valor": START_DATE},
                {"Parâmetro": "Data Final", "Valor": END_DATE},
                {
                    "Parâmetro": "Situações consideradas (A/B/C)",
                    "Valor": "VENDA SEPD, PEDIDO DE VENDA, VDA HOMOLOG",
                },
            ]
        )
        params.to_excel(writer, index=False, sheet_name=sheet, startrow=0)
        ws = writer.sheets[sheet]
        ws.freeze_panes = "A5"

        # Detalhe Diário
        startrow = len(params) + 2
        ws.cell(row=startrow, column=1, value="Detalhe Diário")
        detalhe.to_excel(writer, index=False, sheet_name=sheet, startrow=startrow)
        header_row_idx = startrow + 1
        excel_number_formats(
            ws,
            header_row_idx,
            detalhe,
            money_cols=[
                "Total Líquido (A)",
                "Vl. Médio Produto (D)",
                "Vl. Médio Pedido (E)",
                "Desconto",
                "Acréscimo",
                "Tot. Bruto",
                "Devoluções Venda",
                "Crédito Cliente",
            ],
            int_cols=["Qtde. Produtos (B)", "Qtde. Pedidos (C)"],
        )

        # Resumo por dia da Semana
        startrow = header_row_idx + len(detalhe) + 3
        ws.cell(row=startrow, column=1, value="Resumo por dia da Semana")
        resumo.to_excel(writer, index=False, sheet_name=sheet, startrow=startrow)
        header_row_idx2 = startrow + 1
        excel_number_formats(
            ws,
            header_row_idx2,
            resumo,
            money_cols=[
                "Total Líquido (A)",
                "Vl. Médio Produto (D)",
                "Vl. Médio Pedido (E)",
                "Desconto",
                "Acréscimo",
                "Tot. Bruto",
                "Devoluções Venda",
                "Crédito Cliente",
            ],
            int_cols=["Qtde. Produtos (B)", "Qtde. Pedidos (C)"],
        )

        # Vendas Total do Período
        startrow = header_row_idx2 + len(resumo) + 3
        ws.cell(row=startrow, column=1, value="Vendas Total do Período")
        venda_total_periodo.to_excel(writer, sheet_name=sheet, startrow=startrow)

        # Totais por Tipo de Pagamento
        startrow = startrow + 4 + len(venda_total_periodo)
        ws.cell(row=startrow, column=1, value="Totais por Tipo de Pagamento")
        tipos_pagto.reset_index(names=["Tipo de Pagamento"]).to_excel(
            writer, index=False, sheet_name=sheet, startrow=startrow
        )
        header_row_idx3 = startrow + 1
        excel_number_formats(
            ws,
            header_row_idx3,
            tipos_pagto.reset_index(),
            money_cols=["Venda Bruta", "Devoluções Venda", "Venda Líquida"],
            int_cols=[],
        )

        # Contadores
        startrow = header_row_idx3 + len(tipos_pagto) + 3
        contadores.to_excel(writer, index=False, sheet_name=sheet, startrow=startrow)
        # Formatar coluna Qtde
        for r in range(startrow + 1, startrow + 1 + len(contadores)):
            ws[f"B{r}"].number_format = "0"

        # Resumo por Vendedor (período)
        startrow = startrow + 3 + len(contadores)
        ws.cell(row=startrow, column=1, value="Resumo por Vendedor")
        vendedores.to_excel(writer, index=False, sheet_name=sheet, startrow=startrow)
        header_row_idx4 = startrow + 1
        excel_number_formats(ws, header_row_idx4, vendedores, money_cols=["Venda Líquida"], int_cols=[])

    print(f"🧾 Excel gerado -> {xlsx_path}")
    print("🏁 Concluído.")